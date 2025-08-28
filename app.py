# ===============================
# app_streamlit.py (Final + Sheet Tidak Hadir â‰¥3 Hari + Bulan Huruf di Surat)
# ===============================
import streamlit as st
import pandas as pd
from datetime import datetime, date
import os
import re
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Konfigurasi upload ---
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- Daftar ID yang di-highlight ---
highlight_ids = {"119","111","112","106","13","18","71","19","148","90","82","142","127"}

# --- Helper ---
def clean_id(id_value):
    if pd.isna(id_value):
        return ""
    id_str = str(id_value).strip()
    if id_str.endswith(".0"):
        id_str = id_str[:-2]
    return id_str

def sort_nicely(l):
    def convert(text):
        return int(text) if text.isdigit() else text.lower()
    def alphanum_key(key):
        return [convert(c) for c in re.split("([0-9]+)", key)]
    return sorted(l, key=alphanum_key)

def highlight_id(val):
    if str(val) in highlight_ids:
        return "background-color: lightgreen; color: black; font-weight: bold;"
    return ""

# --- Streamlit UI ---
st.set_page_config(page_title="Rekap Absensi PT. QUANTUM", layout="wide")
st.title("ðŸ“‘ Rekap Absensi PT. QUANTUM")

uploaded_file = st.file_uploader("Upload File Absensi (.xlsx/.xls)", type=["xlsx", "xls"])

if uploaded_file:
    # Simpan sementara
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{uploaded_file.name}")
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    # === BACA SEMUA SHEET ===
    all_sheets = pd.read_excel(file_path, sheet_name=None)

    # Gabung semua sheet
    df_all = []
    for sheet_name, df in all_sheets.items():
        max_cols = len(df.columns)
        column_names = ["Perusahaan","Nama","ID","Tgl/Waktu","Mesin_ID","Kolom6","Status","Kolom8"]
        column_mapping = {}
        for i in range(min(max_cols, len(column_names))):
            column_mapping[column_names[i]] = df.iloc[:, i]
        df_fix = pd.DataFrame(column_mapping)
        df_all.append(df_fix)
    df_fix = pd.concat(df_all, ignore_index=True)

    # Normalisasi
    df_fix["Nama"] = df_fix["Nama"].astype(str).str.strip()
    df_fix["ID"] = df_fix["ID"].apply(clean_id)
    df_fix = df_fix[df_fix["Nama"].notna() & (df_fix["Nama"]!="nan") & (df_fix["Nama"]!="")]
    df_fix = df_fix[df_fix["ID"].notna() & (df_fix["ID"]!="nan") & (df_fix["ID"]!="")]
    df_fix["Tgl/Waktu"] = pd.to_datetime(df_fix["Tgl/Waktu"], dayfirst=True, errors="coerce")
    df_fix = df_fix.dropna(subset=["Tgl/Waktu"])
    df_fix["Tanggal_Saja"] = df_fix["Tgl/Waktu"].dt.date
    df_fix = df_fix.drop_duplicates(subset=["ID","Tanggal_Saja"])

    # ID unik
    semua_id_unik = sort_nicely(list(set(df_fix["ID"])))

    # Rentang tanggal kerja (tanpa Minggu)
    if not df_fix["Tgl/Waktu"].empty:
        tanggal_awal = df_fix["Tgl/Waktu"].dt.date.min()
        tanggal_akhir = df_fix["Tgl/Waktu"].dt.date.max()
        semua_tanggal = [tgl for tgl in pd.date_range(tanggal_awal, tanggal_akhir).date if pd.Timestamp(tgl).weekday()!=6]
    else:
        semua_tanggal = []

    # Mapping ID â†’ Nama
    id_to_nama = dict(zip(df_fix["ID"], df_fix["Nama"]))

    # ========================
    # REKAP TELAT
    # ========================
    jam_telat = datetime.strptime("07:50:00", "%H:%M:%S").time()
    df_pagi = df_fix[(df_fix["Tgl/Waktu"].dt.hour >=5) & (df_fix["Tgl/Waktu"].dt.hour <=9)]
    rekap_telat = []
    for id_karyawan in semua_id_unik:
        nama_karyawan = id_to_nama.get(id_karyawan, "Unknown")
        data_id = df_pagi[df_pagi["ID"]==id_karyawan]
        telat_id = data_id[data_id["Tgl/Waktu"].dt.time > jam_telat]
        for _, row in telat_id.iterrows():
            rekap_telat.append({
                "ID": id_karyawan,
                "Nama": nama_karyawan,
                "Tgl/Waktu Telat": row["Tgl/Waktu"]
            })
    df_telat = pd.DataFrame(rekap_telat)

    # ========================
    # REKAP TIDAK HADIR + JUMLAH
    # ========================
    rekap_tidak_hadir = []
    jumlah_absen_total = []
    for id_karyawan in semua_id_unik:
        nama_karyawan = id_to_nama.get(id_karyawan,"Unknown")
        data_id = df_fix[df_fix["ID"]==id_karyawan]
        hadir_tanggal = set(data_id["Tgl/Waktu"].dt.date) if not data_id.empty else set()
        tidak_hadir_tanggal = [tgl for tgl in semua_tanggal if tgl not in hadir_tanggal]
        for tgl in tidak_hadir_tanggal:
            rekap_tidak_hadir.append({"ID":id_karyawan,"Nama":nama_karyawan,"Tanggal Tidak Hadir":tgl})
        jumlah_absen_total.append({
            "ID":id_karyawan,
            "Nama":nama_karyawan,
            "Jumlah Absen Total":len(hadir_tanggal)
        })
    df_tidak_hadir = pd.DataFrame(rekap_tidak_hadir)
    df_jumlah_absen = pd.DataFrame(jumlah_absen_total)

    if not df_tidak_hadir.empty:
        jumlah_tidak_hadir = df_tidak_hadir.groupby("ID").size().reset_index(name="Jumlah Tidak Hadir")
        df_jumlah_absen = pd.merge(df_jumlah_absen, jumlah_tidak_hadir, on="ID", how="left")
    else:
        df_jumlah_absen["Jumlah Tidak Hadir"] = 0
    df_jumlah_absen["Jumlah Tidak Hadir"] = df_jumlah_absen["Jumlah Tidak Hadir"].fillna(0).astype(int)

    # --- TAMPILKAN DI STREAMLIT ---
    st.subheader("ðŸ“Œ Rekap Telat")
    if not df_telat.empty:
        st.dataframe(df_telat.style.applymap(highlight_id, subset=["ID"]))
    else:
        st.info("Tidak ada data karyawan telat")

    st.subheader("ðŸ“Œ Rekap Tidak Hadir")
    if not df_tidak_hadir.empty:
        st.dataframe(df_tidak_hadir.style.applymap(highlight_id, subset=["ID"]))
    else:
        st.info("Tidak ada data karyawan tidak hadir")

    st.subheader("ðŸ“Œ Jumlah Kehadiran")
    st.dataframe(df_jumlah_absen.style.applymap(highlight_id, subset=["ID"]))

    # === Simpan semua sheet utama ke Excel ===
    hasil_rekap_path = os.path.join(UPLOAD_FOLDER, f"hasil_rekap_{uploaded_file.name}")
    with pd.ExcelWriter(hasil_rekap_path, engine="openpyxl") as writer:
        if not df_telat.empty:
            df_telat.to_excel(writer, sheet_name="Karyawan Telat", index=False)
        if not df_tidak_hadir.empty:
            df_tidak_hadir.to_excel(writer, sheet_name="Karyawan Tidak Hadir", index=False)
        df_jumlah_absen.to_excel(writer, sheet_name="Jumlah Kehadiran", index=False)

    # --- Load workbook untuk highlight & tambah sheet baru ---
    wb = load_workbook(hasil_rekap_path)

    # Highlight ID tertentu di semua sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if str(cell.value) in highlight_ids:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    cell.font = Font(color="000000", bold=True)

    # --- Sheet Tidak Hadir â‰¥3 Hari ---
    df_tidak_hadir_lebih3 = df_jumlah_absen[df_jumlah_absen["Jumlah Tidak Hadir"]>=3].copy()
    if not df_tidak_hadir_lebih3.empty:
        ws2 = wb.create_sheet(title="Tidak Hadir â‰¥3 Hari")
        for r_idx, row in enumerate(dataframe_to_rows(df_tidak_hadir_lebih3, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        # Highlight ID tertentu di sheet baru
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=1):
            for cell in row:
                if str(cell.value) in highlight_ids:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    cell.font = Font(color="000000", bold=True)

    wb.save(hasil_rekap_path)

    # --- Download file Excel ---
    with open(hasil_rekap_path, "rb") as f:
        st.download_button("ðŸ“¥ Download Rekap Excel", f, file_name=os.path.basename(hasil_rekap_path))

    # --- Surat Panggilan (â‰¥3 Tidak Hadir) ---
    if not df_tidak_hadir_lebih3.empty:
        st.subheader("ðŸ“Œ Surat Panggilan (â‰¥3 Tidak Hadir)")
        df_tidak_hadir_lebih3 = df_tidak_hadir_lebih3.sort_values(by="Jumlah Tidak Hadir", ascending=False)
        st.dataframe(df_tidak_hadir_lebih3[["ID","Nama","Jumlah Tidak Hadir"]]
                     .style.applymap(highlight_id, subset=["ID"]))

        hari_list = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]
        for _, row in df_tidak_hadir_lebih3.iterrows():
            spg_filename = f"surat_panggilan_{row['ID']}_{uploaded_file.name.rsplit('.',1)[0]}.docx"
            spg_path = os.path.join(UPLOAD_FOLDER, spg_filename)
            template_path = os.path.join("templates", "template_surat_panggilan.docx")
            doc = DocxTemplate(template_path)

            df_absen_id = df_tidak_hadir[df_tidak_hadir["ID"]==row['ID']]
            semua_tgl = df_absen_id["Tanggal Tidak Hadir"].apply(lambda x: x.strftime("%d %B %Y")).tolist()
            tanggal_terakhir = ", ".join(semua_tgl)
            jumlah_hari = len(semua_tgl)

            tanggal_surat = date.today()
            nama_hari = hari_list[tanggal_surat.weekday()]

            context = {
                "NAMA": row["Nama"],
                "ID": row["ID"],
                "JUMLAH_HARI": jumlah_hari,
                "TANGGAL_ABSEN": tanggal_terakhir,
                "TANGGAL_SURAT": f"{nama_hari}, {tanggal_surat.strftime('%d %B %Y')}"
            }

            doc.render(context)
            doc.save(spg_path)

            with open(spg_path, "rb") as f:
                st.download_button(f"ðŸ“¥ Download Surat Panggilan untuk {row['Nama']}", f, file_name=spg_filename)

